from asyncio.windows_events import NULL
from cmath import log
import configparser
import os
import logging
from datetime import datetime, timedelta
from re import S
from tkinter.tix import COLUMN
from dateutil.relativedelta import relativedelta
import pandas as pd
import psycopg2
import pymssql
import openpyxl
import copy
from openpyxl.styles import Alignment,Font,Border,Side,PatternFill
from openpyxl.utils import get_column_letter
import sys

'''-----------------------------讀取ini檔 並 包進 Dict格式------------------------------'''
def load_conf(config_file) : 

    def null_proc(text) :
        if text in ["None","Null","none","null",""] :
            return "None"
        else : return text

    try : 
        iniConfig = {}
        iniFile = configparser.ConfigParser()
        iniFile.read(config_file,encoding = "utf-8")
        cfgSections = iniFile.sections()
        for section in cfgSections : 
            iniConfig[section] = {}
            cfgOptions = iniFile.options(section)
            for op in cfgOptions :
                iniConfig[section][op] = null_proc(iniFile.get(section,op))
        return iniConfig
        
    except Exception as e : 
        logging.error('Info','Read Report.ini file error')
        logging.exception(e)
'''---------------------AutoRun : 判斷執行時間進行該時間應該被產出檔案---------------------------------'''
def runevent() :

    runReport = []
    timestamp = datetime.now()
    
    if timestamp.strftime("%H%M") == "0001" :
        runReport.append("DayStart")
        if timestamp.strftime("%w") == "1" :
            runReport.append("WeekStart")
        if timestamp.strftime("%d") == "01" :
            runReport.append("MonthStart") 
            if timestamp.strftime("%m") == "01" :
                runReport.append("YearStart")
    else :
        runReport.append("None")

    return runReport

#-------------------------------Report Process--------------------------------------
                   
# 定義 工作資料夾及 Config 檔案路徑
#workPath = os.path.dirname(os.path.realpath(sys.executable))

workPath = os.path.dirname(sys.executable)
#workPath = os.path.dirname(__file__)
workConfig = workPath + "\Config.ini"
'''---------------------定義 logging 格式---------------------------------'''
logging.basicConfig(level=logging.DEBUG,
                    format="%(asctime)s %(levelname)s - %(message)s",
                    datefmt= "%Y/%m/%d %H:%M:%S",
                    handlers=[logging.FileHandler(workPath + "\\" + "Report_log.txt","a","utf-8"),])

# 判斷Config 檔案是否存在，如果是 將各個設定檔讀出，如果否 則報錯
if os.path.isfile(workConfig) :

    iniFile = load_conf(workConfig)
    # Basic : (1) 從 Option 字串中帶有"rptfilename" 的值 定義 產出檔案名稱資訊，並自動加入產出時間戳記
    #         以及 分別定義 該檔案從 Template 中要讀取的 SheetName，並針對個別 Sheet 定義撈取的時間單位及區間
    #         (2) 定義產出報表檔案的路徑
    
    rptSqlExcu = []
    itemCount = 0
    for item in iniFile.get("Basic").items() : 
        if "rptfi1ename" in item[0] and datetime.now().strftime("%Y") == "2022" : 
            rptSqlExcu.append((lambda x : x.split(".",9))(item[0]))
            rptSqlExcu[itemCount].append("{}_{}.xlsx".format(item[1],datetime.now().strftime("%Y%m%d%H%M%S")))
            itemCount += 1 
        elif "rptfilename" in item[0] :           
            rptSqlExcu.append((lambda x : x.split(".",9))(item[0]))
            rptSqlExcu[itemCount].append("{}_{}.xlsx".format(item[1],datetime.now().strftime("%Y%m%d%H%M%S")))
            itemCount += 1
    rptFilePath = iniFile.get("Basic").get("rptfilepath")
    
    logging.getLogger().debug(rptSqlExcu)

    # DatabaseConfig : 資料庫基本設定
    myDatabase = iniFile.get("DatabaseConfig").get("database")
    myDatabaseSource = iniFile.get("DatabaseConfig").get("databasesource")
    myDatabaseIP = iniFile.get("DatabaseConfig").get("databaseip")
    myDatabasePort = iniFile.get("DatabaseConfig").get("databaseport")
    myDatabaseUser = iniFile.get("DatabaseConfig").get("databaseuser")
    myDatabasePassword = iniFile.get("DatabaseConfig").get("databasepassword")

    # Target -- 統整各個[0]FileName 對應的 [1]SheetName 及其 撈取SQL的參數值 : [2]時間單位、[3]時間區間、[4]TableName
    for item in iniFile.get("Target").items() :
        for i in range(len(rptSqlExcu)) :
            if item[0] in rptSqlExcu[i] :
                rptSqlExcu[i][0] = item[1]
                rptSqlExcu[i].reverse()
    logging.getLogger().debug(rptSqlExcu)

    # ReportTemplate : Tempalte 的名稱
    tmpFileName = iniFile.get("ReportTemplate").get("tmpfilename")
    
else :
    logging.getLogger().info("info","Report config file not found")

# 資料庫連線
try :

    conSQLsts = False
    if myDatabase == "PostgreSQL" :
        conSQL = psycopg2.connect(database = myDatabaseSource,
                                    user = myDatabaseUser,
                                    password = myDatabasePassword,
                                    host = myDatabaseIP,
                                    port = myDatabasePort )
        cur = conSQL.cursor()
        conSQLsts = True

    elif myDatabase == "MSSQL" :
        conSQL = pymssql.connect(server = myDatabaseIP,
                                    user = myDatabaseUser,
                                    password = myDatabasePassword,
                                    database = myDatabaseSource)
        cur = conSQL.cursor()
        conSQLsts = True
except Exception as e : 
    logging.getLogger().info("SQL Connect Error")
    logging.getLogger().exception(e)

# 讀取 Report 樣板  (限制 : 需與執行檔同一個資料夾)
if os.path.isfile(workPath + "\\" + tmpFileName) : 
    rptTemplate = pd.read_excel(workPath + "\\" + tmpFileName,sheet_name = None,header = None)
else : 
    logging.getLogger().info("Can not find the Report Template")

# 針對 個別 Sheet 指定的 Table 撈取資料，並將其打入 Excel 另存新檔
if conSQLsts == True :
    # rptSQLExcute : [0]FileName 對應的 [1]SheetName 及其 撈取SQL的參數值 : [2]時間單位、[3]時間區間、[4]TableName
    # sheet[0] : Template 內的 SheetName ; sheet[1].loc[3]為使用者自key撈取欄位名稱或時間轉換(根據 SQL 不同改變)
    for sheet in rptTemplate.items() :
        for i in range(len(rptSqlExcu)) : 
            if sheet[0].lower() in rptSqlExcu[i] : 
                
                rptSqlExcu[i].append(
                    [x.replace("[","").replace("]","")
                    for x in sheet[1].loc[3].tolist() 
                    if pd.isnull(x) == False and x != 'nan'])
                
                # rptSQLExcute : [0]FileName 對應的 [1]SheetName 及其 撈取SQL的參數值 : [2]時間單位、[3]時間區間、[4]TableName [5] SELECT、FROM、WHERE
                # WHERE = [0] [1] 是時間戳記 [2] ... 以後為撈取欄位
                try : 
                    
                    if myDatabase == "PostgreSQL" :
                        # interval '單位' '區間'
                        
                        cur.execute("SELECT " + ",".join(rptSqlExcu[i][5])
                            + " FROM " + rptSqlExcu[i][4]
                            + " WHERE " + (lambda x : x[x.index("as")+3:])(rptSqlExcu[i][5][1]) + 
                            " >= cast(to_char(now(),'YYYY/MM/DD HH24:MI') as timestamp with time zone) - interval '" +
                            rptSqlExcu[i][3] + " " + rptSqlExcu[i][2] + "' order by " + (lambda x : x[x.index("as")+3:])(rptSqlExcu[i][5][0])
                            + " asc")
                        logging.getLogger().debug("SELECT " + ",".join(rptSqlExcu[i][5])
                            + " FROM " + rptSqlExcu[i][4]
                            + " WHERE " + (lambda x : x[x.index("as")+3:])(rptSqlExcu[i][5][1]) + 
                            " >= cast(to_char(now(),'YYYY/MM/DD HH24:MI') as timestamp with time zone) - interval '" +
                            rptSqlExcu[i][3] + " " + rptSqlExcu[i][2] + "' order by " + (lambda x : x[x.index("as")+3:])(rptSqlExcu[i][5][0])
                            + " asc")
                        rptData = pd.DataFrame(cur.fetchall())
                        
                    
                    elif myDatabase == "MSSQL" :
                        # DATEADD '區間' '單位'
                        cur.execute("SELECT " + ",".join(rptSqlExcu[i][5])
                            + " FROM " + myDatabaseSource + "." + rptSqlExcu[i][4]
                            + " WHERE " + (lambda x : x[x.index("as")+3:])(rptSqlExcu[i][5][1]) +
                            " >= DATEADD(" + rptSqlExcu[i][2] + ", -" + rptSqlExcu[i][3] + 
                            ",format(getdate(),'yyyy/MM/dd hh:mm')) order by " + (lambda x : x[x.index("as")+3:])(rptSqlExcu[i][5][0])
                            + " asc")
                        logging.getLogger().debug("SELECT " + ",".join(rptSqlExcu[i][5])
                            + " FROM " + myDatabaseSource + "." + rptSqlExcu[i][4]
                            + " WHERE " + (lambda x : x[x.index("as")+3:])(rptSqlExcu[i][5][1]) +
                            " >= DATEADD(" + rptSqlExcu[i][2] + ", -" + rptSqlExcu[i][3] + 
                            ",format(getdate(),'yyyy/MM/dd hh:mm')) order by " + (lambda x : x[x.index("as")+3:])(rptSqlExcu[i][5][0])
                            + " asc")
                        
                        rptData = pd.DataFrame(cur.fetchall())        

                except Exception as e : 
                    logging.getLogger().info('Read Report.ini file error')
                    logging.getLogger().exception(e)
                
                pd.concat([sheet[1],rptData],axis = 0,ignore_index = True).drop([3]).to_excel(rptFilePath + "\\" + rptSqlExcu[i]
                [0],index = False,header = False)

    conSQL.close()

#資料表美觀
try : 
    wb_tmp = openpyxl.load_workbook(workPath + "\\" + tmpFileName)

    for sheetName in wb_tmp.sheetnames : 
        wb_tmp[sheetName].delete_rows(4)
        rptName_num = 0
        for report in rptSqlExcu :
            
            if sheetName.lower() == report[1] :
                
                wb_rpt = openpyxl.load_workbook(rptFilePath + "\\" + report[0])
                # Sheet Tab 顏色同步
                wb_rpt["Sheet1"].sheet_properties.tabColor = wb_tmp[sheetName].sheet_properties.tabColor

                # 欄位格式 同步
                wb_tmp_merge = list(wb_tmp[sheetName].merged_cells)
                if len(wb_tmp_merge) > 0 :
                    for i in range(0 , len(wb_tmp_merge)) :

                        wb_rpt["Sheet1"].merge_cells((lambda x : x.replace('(<CellRange ', '').replace('>,)', ''))(str(wb_tmp_merge[i])))
                
                for i , row in enumerate(wb_tmp[sheetName].iter_rows()) :
                    wb_rpt["Sheet1"].row_dimensions[i+1].height = wb_tmp[sheetName].row_dimensions[i+1].height
                    for j , cell in enumerate(row) :
                        wb_rpt["Sheet1"].column_dimensions[get_column_letter(j+1)].width = wb_tmp[sheetName].column_dimensions[get_column_letter(j+1)].width
                        #wb_rpt["Sheet1"].cell(row = i + 1, column = j + 1 , value = cell.value)

                        source_cell = wb_tmp[sheetName].cell(i + 1 , j + 1)
                        target_cell = wb_rpt["Sheet1"].cell(i + 1 , j + 1)
                        target_cell.fill = copy.copy(source_cell.fill)
                        
                        if source_cell.has_style : 
                            target_cell._style = copy.copy(source_cell._style)
                            target_cell.font = copy.copy(source_cell.font)
                            target_cell.border = copy.copy(source_cell.border)
                            target_cell.fill = copy.copy(source_cell.fill)
                            target_cell.number_format = copy.copy(source_cell.number_format)
                            target_cell.protection = copy.copy(source_cell.protection)
                            target_cell.alignment = copy.copy(source_cell.alignment)
                        
                        if target_cell.value == "開始時間" : 
                            if report[2] == "hour" :
                                wb_rpt["Sheet1"].cell(i + 1 , j + 2).value = (datetime.now() + timedelta(hours=int(report[3])*(-1))).strftime("%Y/%m/%d %H:%M")
                            elif report[2] == "week" : 
                                wb_rpt["Sheet1"].cell(i + 1 , j + 2).value = (datetime.now() + relativedelta(weeks=int(report[3])*(-1))).strftime("%Y/%m/%d %H:%M")
                            elif report[2] == "month" : 
                                wb_rpt["Sheet1"].cell(i + 1 , j + 2).value = (datetime.now() + relativedelta(months=int(report[3])*(-1))).strftime("%Y/%m/%d %H:%M")
                            elif report[2] == "year" : 
                                wb_rpt["Sheet1"].cell(i + 1 , j + 2).value = (datetime.now() + relativedelta(years=int(report[3])*(-1))).strftime("%Y/%m/%d %H:%M")  
                        elif target_cell.value == "結束時間" : 
                            wb_rpt["Sheet1"].cell(i + 1 , j + 2).value = datetime.now().strftime("%Y/%m/%d %H:%M")
                        elif target_cell.value == "製表日期" : 
                            wb_rpt["Sheet1"].cell(i + 2 , j + 1).value = datetime.now().strftime("%Y/%m/%d %H:%M:%S")

                for i , row in enumerate(wb_rpt["Sheet1"].iter_rows()) :    
                    for j , cell in enumerate(row) :    
                        if i > 2 and wb_rpt["Sheet1"].cell(i + 1 , 1).value != None :
                            wb_rpt["Sheet1"].cell(i + 1 , j + 1).border = Border(top = Side(border_style = "thin" , color = "FF000000"),
                                                                                right = Side(border_style = "thin" , color = "FF000000"),
                                                                                bottom = Side(border_style = "thin" , color = "FF000000"),                                                                                
                                                                                left = Side(border_style = "thin" , color = "FF000000"))
                            wb_rpt["Sheet1"].cell(i + 1 , j + 1).font = Font(size = 12)
                                                                                                            
                         
                wb_rpt.save(rptFilePath + "\\" + report[0])
                wb_rpt.close()
                
            rptName_num += 1
    wb_tmp.close()
except Exception as e : 
    logging.getLogger().info('Excel modify Error')
    logging.getLogger().exception(e)

logging.getLogger().info("Report Process Done")

#rptoldPath = rptFilePath + "\\" + rptFileName[i]


